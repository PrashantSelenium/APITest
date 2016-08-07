'''
@This file is used to read from the excels, gather the info and do the appropriate parsing.
@Author : Runjhun Singh
'''
import openpyxl
import openpyxl.cell
from openpyxl.utils import (_get_column_letter)

import os, sys
from openpyxl.styles import colors , Font, Color , Style
from openpyxl.styles import Alignment
import ExecuteTestCase
import re
import json
import time
from random import choice
from string import ascii_uppercase
from bs4 import BeautifulSoup
from collections import namedtuple
from argparse import Namespace

# global method

currentFilePath = os.path.abspath(__file__)	
rootPath = "\\".join(currentFilePath.split("\\")[:-2])
sys.path.append(rootPath)
from utils.TestCaseValues import TestCaseValues

def getRootPath():
	currentFilePath = os.path.abspath(__file__)	
	rootPath = "\\".join(currentFilePath.split("\\")[:-2])
	return rootPath

def getExcelFilePath(filename):
	# print "--finding path of the excel file----"
	# print rootPath
	rootPath = getRootPath()
	input_file_path = rootPath+filename
	return input_file_path

def openWorkbook(filename):
	global wb
	input_file_path = getExcelFilePath(filename)
	wb = openpyxl.load_workbook(input_file_path)
	return wb

def openExcelSheet(wb,sheet_name):
	# print wb
	sheet = wb.get_sheet_by_name(sheet_name)
	return sheet

def getTotalTestCases(workbook):
	# input_file_path = getExcelFilePath(filename)
	sheet = openExcelSheet(workbook,"TestSuite1")
	row = sheet.max_row
	return row , sheet

def readFromExcelSheet(row,sheet):
	if (sheet['A' + str(row)].value != ""):
		testCaseValues = TestCaseValues(sheet,row)
		print testCaseValues.test
		# print testCaseValues.request_body
		testCaseValues.request_body = createRequestBody(testCaseValues.request_body)	
		global method
		method = testCaseValues.method
		print "after random generator"  + testCaseValues.request_body
		return testCaseValues

def createRequestBody(request):
	# print request
	request = re.sub(r'#\w+', replaceData, request)
	return request

def replaceData(match):
	# print "in replaceData"
	match = match.group()
	if match == "#random_string":
		matching_value = getRandomSring()
		print matching_value
		# if (method == "SOAP"):
		# 	return   matching_value 
		# else:
		return  "\"" + matching_value + "\""

	else:
		matching_value = getTestCaseValuesFromExcel(match)
		# print matching_value
		return  matching_value

def getTestCaseValuesFromExcel(variable):
	# print variable
	# input_file_path = getExcelFilePath("\TestCase Repository.xlsx")
	sheet = openExcelSheet(wb,"TestData")
	counter = 0
	for column in range(1,20):
	    column_letter = _get_column_letter(column)
	    for row in range(1,sheet.max_row):
	        if sheet[column_letter + str(row)].value == variable:
		        cell = sheet[column_letter + str(row+1)].value

		        if (method == "SOAP"):
		        	return   str(cell) 
		        else:
		        	return "\"" + str(cell) + "\""

def getRandomSring():
	print "in random string"
	return ''.join(choice(ascii_uppercase) for i in range(12))


def writeIntoExcelSheet(responseJson,sheet,row,teststatus,api_type):
	a1 = sheet['L' + str(row)]
	if teststatus:
		ft = Font(color=colors.BLACK)
		status = "Pass"
	else :
		ft = Font(color=colors.RED)
		status = "Fail"
		
	a1.font = ft
	if api_type:
		sheet['L' + str(row)].alignment = Alignment(wrapText=True)
		sheet['L' + str(row)] = str(json.dumps(responseJson, indent=4))
		sheet['M' + str(row)] = str(json.dumps(status))
	else :
		sheet['L' + str(row)].alignment = Alignment(wrapText=True)
		sheet['L' + str(row)] = BeautifulSoup(responseJson, "xml").prettify()
		sheet['M' + str(row)] = str(json.dumps(status))

# def writeintoexcelTestData(filename,variable):
# 	wrkbk = openWorkbook(filename)
# 	print " in writeintoexcelTestData"
# 	sheet = openExcelSheet(wrkbk,"TestData")
# 	counter = 0
# 	cell = 1234
# 	print method	
# 	for column in range(1,20):
		
# 	    column_letter = _get_column_letter(column)
	   
# 	    for row in range(1,sheet.max_row+1):
# 	    	print "in for  loop"
# 	    	print sheet[column_letter + str(row)].value
# 	    	print variable
# 	        if sheet[column_letter + str(row)].value == variable:
# 	        	print " first if"
		        
# 		        if (method == "SOAP"):
# 		        	print "in SOAP"
# 		        	# sheet['A2'].value = "Testing"
# 		        	sheet[column_letter + str(row+1)].value  = cell
# 		        	savelocation = getRootPath()
# 		        	print "saving the file"
# 		        	wrkbk.save(savelocation + filename)
# 		        else:
# 		        	sheet[column_letter + str(row+1)].value  = "\"" + cell + "\""
				
# 		return

def fetchValuesToWritefromResponse(response):
	print "in fetchValuesToWritefromResponse"
	# print response
	print response['ns0:Envelope']['ns0:Body']['ns1:LoginResponse']['return']['ResponseMessage']

def saveResults(workbook,filename):
	savelocation = getRootPath()
	print savelocation
workbook.save( savelocation + '\TestExecution' + str(time.strftime("%Y%m%d-%H%M%S")) + '.xlsx')